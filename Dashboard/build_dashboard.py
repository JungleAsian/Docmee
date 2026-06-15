# -*- coding: utf-8 -*-
"""
Docmee Command Center — dashboard generator.

Reads the source-of-truth tracker (docmee-trackers.xlsx) + monorepo git state and
emits ONE self-contained HTML file (Docmee-Dashboard.html) with all data, styles,
and hand-rolled SVG charts inlined — opens offline by double-click, no server, no CDN.

Re-run this whenever the tracker or repo changes:  python build_dashboard.py
"""
import json, subprocess, datetime, pathlib, sys
from collections import Counter, defaultdict

ROOT     = pathlib.Path(__file__).resolve().parent.parent      # ...\Docmee
XLSX     = ROOT / "Trackers" / "docmee-trackers.xlsx"
MONOREPO = ROOT / "monorepo"
OUT      = pathlib.Path(__file__).resolve().parent / "Docmee-Dashboard.html"
TEMPLATE = pathlib.Path(__file__).resolve().parent / "template.html"

def load_wb():
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl missing — run: pip install openpyxl")
    return openpyxl.load_workbook(XLSX, data_only=True)

def cell(v):
    return "" if v is None else str(v).strip()

def sheet_rows(ws, min_row=1):
    out = []
    for r in ws.iter_rows(min_row=min_row, values_only=True):
        vals = [cell(c) for c in r]
        if any(vals):
            out.append(vals)
    return out

# ----- agent lane model (from docmee-dev-plan.md §4) --------------------------
PHASES = [
    ("S0","Sprint 0","Setup","Setup & seam",0,1,
     "Infra, CI, RLS test harness, SEC06 decision, contract v0",
     "App shell, auth flow, design system, i18n framework, mock server",
     "Auth round-trips against the real API"),
    ("0","Phase 0","Foundation","Foundation",0,2,
     "RLS, encryption, chokepoint, idempotency, deploy",
     "Login, clinic-switch, settings scaffold (mocked)",
     "Tenant-scoped session works end-to-end"),
    ("1A","Phase 1A","MVP","Core Bot",2,6,
     "KB, 5-intent pipeline, six-gate, transcription",
     "KB editor UI, inbox shell (mocked), bot-mode toggle",
     "KB CRUD live; a message appears in the inbox"),
    ("1B","Phase 1B","MVP","Human Inbox",6,9,
     "Patient/conversation/notes APIs, assignment, capture allowlist",
     "Unified inbox, CRM, assignment, notes, tags",
     "Full inbox operates on real data"),
    ("1C","Phase 1C","MVP","Scheduling",9,12,
     "Calendar truth, intake state machine, lifecycle",
     "Booking/calendar UI, appointment views",
     "Book -> appears on calendar + panel"),
    ("GATE","Pilot Launch Gate","Gate","Pilot Gate",12,12,
     "Rate-limiting (SEC18), backups (X12)",
     "Polish, empty/error states, a11y pass",
     "Launch-gate checklist green"),
    ("2A","Phase 2A","Pro ops","Multi-User",12,14,
     "RBAC, routing, notifications, invoicing",
     "Role mgmt, IA Studio admin, panel i18n (ES/EN), quick replies",
     "RBAC enforced in UI + API"),
    ("2B","Phase 2B","Pro ops","Channels",14,16,
     "Messenger/IG adapters, unified pipeline",
     "Channel badges, connect flow, merge UI",
     "Cross-channel message in one inbox"),
    ("2C","Phase 2C","Pro ops","Automation",16,18,
     "Templates, six-gate jobs, reminders",
     "Template manager, automation settings UI",
     "Reminder fires through the six-gate"),
    ("2D","Phase 2D","Pro ops","Analytics",18,20,
     "Rollups, error-review, full-text search",
     "Dashboards, error-review UI, search",
     "Charts render from real rollups"),
    ("3A","Phase 3A","Scale","Multi-Doctor",20,22,
     "Doctor entity, per-doctor calendar/KB",
     "Doctor mgmt, doctor-select in booking",
     "Book a specific doctor"),
    ("3B","Phase 3B","Scale","Adv. Intake",22,23,
     "Flow engine, rule engine, copilot API",
     "Flow builder UI, rule editor, copilot panel",
     "A custom flow runs end-to-end"),
    ("3C","Phase 3C","Scale","Integrations",23,25,
     "OCR, Sheets/CRM export, reports",
     "Doc-upload UI, export config, report views",
     "OCR'd doc enters KB; export with consent"),
    ("3D","Phase 3D","Scale","Mobile/PWA",25,26,
     "Push service (VAPID)",
     "PWA manifest/SW, push opt-in, responsive pass",
     "Installable PWA receives a push"),
]

# Status reflects on-disk reality as of last refresh (code built locally, largely
# uncommitted). Statuses are user-editable in the dashboard (saved per browser).
SPRINT0_TICKETS = [
    ("S0-P1","Prime","Finalize contract v0","Confirm /auth/session + Phase-0/1 shapes in packages/contracts; generate TS types (pnpm contract:types).","done"),
    ("S0-P2","Prime","API + RLS test harness","Bootstrap apps/api (Fastify), wire Supabase + auth middleware, implement GET /auth/session, create RLS/tenant-isolation test scaffold.","done"),
    ("S0-A1","Alpha","App shell + design system","Bootstrap apps/web (Next.js 14), design system in packages/ui, i18n (ES/EN), run mock server from the contract (pnpm contract:mock).","done"),
    ("S0-A2","Alpha","Auth/login against mock","Build login + clinic-context flow against the mocked /auth/session; switches to real at the checkpoint.","doing"),
]

# Per-phase track seed: agent -> phase-id -> status. Default "todo"; overrides here.
TRACK_SEED = {
    "prime:S0": "done",    # backend Sprint-0 built & green
    "alpha:S0": "doing",   # FE scaffold built; checkpoint (flip to real API) not yet green
}

PRIME_PKGS = [
    ("apps/api","Fastify backend"),
    ("apps/worker","BullMQ workers"),
    ("packages/contracts","OpenAPI seam (owner)"),
    ("packages/db","Supabase schema, RLS"),
    ("packages/core","Pipeline, six-gate, rules"),
    ("packages/channels","Meta/WA/Messenger adapters"),
]
ALPHA_PKGS = [
    ("apps/web","Next.js 14 panel"),
    ("packages/ui","Design system"),
]

def build_data():
    wb = load_wb()

    # ---- Phase Gates: pass/pending per phase --------------------------------
    gate_by_phase = defaultdict(lambda: Counter())
    gate_rows = []
    ws = wb["Phase Gates"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        ph, num, crit, result, notes = (list(r) + [None]*5)[:5]
        ph = cell(ph)
        if not ph or ph.lower() == "phase":
            continue
        result = cell(result) or "Pending"
        gate_by_phase[ph][result] += 1
        gate_rows.append({"phase": ph, "n": cell(num), "criterion": cell(crit),
                          "result": result, "notes": cell(notes)})

    # ---- Action Tracker -----------------------------------------------------
    actions, action_status = [], Counter()
    for r in sheet_rows(wb["Action Tracker"], 2):
        if r and r[0].startswith("X"):
            rid, item, owner, blocks, lead, status = (r + [""]*6)[:6]
            actions.append({"id": rid, "item": item, "owner": owner,
                            "blocks": blocks, "lead": lead, "status": status})
            action_status[status or "Unknown"] += 1

    # ---- Security Audit -----------------------------------------------------
    security, sec_priority, sec_status_simple = [], Counter(), Counter()
    def simplify_sec(s):
        s = s.lower()
        if s.startswith("mitigated") or s.startswith("locked") or s.startswith("closed"):
            return "Mitigated"
        if "open" in s or "partial" in s:
            return "Open"
        return "Other"
    for r in sheet_rows(wb["Security Audit"], 2):
        if r and r[0].startswith("SEC"):
            sid, cat, concern, details, blast, prio, sol, status = (r + [""]*8)[:8]
            security.append({"id": sid, "cat": cat, "concern": concern,
                             "blast": blast, "priority": prio, "status": status})
            sec_priority[prio or "?"] += 1
            sec_status_simple[simplify_sec(status)] += 1

    # ---- Risk Register ------------------------------------------------------
    risks, risk_priority = [], Counter()
    for r in sheet_rows(wb["Risk Register"], 2):
        if r and r[0].startswith("R") and r[0][1:2].isdigit():
            rid, risk, owner, affects, like, sev, prio, path, disp = (r + [""]*9)[:9]
            risks.append({"id": rid, "risk": risk, "owner": owner, "likelihood": like,
                          "severity": sev, "priority": prio, "status": disp})
            risk_priority[prio or "?"] += 1

    # ---- Decisions Index ----------------------------------------------------
    decisions_by_phase = Counter()
    decisions_total = 0
    flagged = 0
    for r in sheet_rows(wb["Decisions Index"], 2):
        if r and r[0] and r[0].lower() != "phase":
            decisions_total += 1
            decisions_by_phase[r[0]] += 1
            if len(r) > 5 and r[5]:
                flagged += 1

    # ---- Gaps ---------------------------------------------------------------
    gaps_by_phase = Counter()
    gaps_status = Counter()
    for r in sheet_rows(wb["Gaps"], 2):
        if r and r[0].startswith("G") and r[0][1:3].isdigit():
            gaps_by_phase[r[1] if len(r) > 1 else "?"] += 1
            gaps_status[r[4] if len(r) > 4 else "?"] += 1

    # ---- Future Improvements ------------------------------------------------
    future = []
    for r in sheet_rows(wb["Future Improvements"], 2):
        if r and r[0].startswith("FI"):
            future.append({"id": r[0], "improvement": r[1] if len(r) > 1 else "",
                           "trigger": r[3] if len(r) > 3 else ""})

    # ---- Cost Tracker -------------------------------------------------------
    def num(v):
        try:
            return round(float(v), 2)
        except (TypeError, ValueError):
            return 0.0
    cost_items, assumptions, cost_totals = [], {}, {}
    for r in wb["Cost Tracker"].iter_rows(values_only=True):
        a = cell(r[0]); b = cell(r[1]) if len(r) > 1 else ""
        if a in ("Dev rate ($/hr)", "AI cost per use ($/session)", "Hours per week (capacity)"):
            assumptions[a] = num(r[1])
        if (a.startswith("P") and a[1:2].isdigit()) or a == "PM":
            cost_items.append({
                "id": a, "name": b, "weeks": num(r[2]), "est_hrs": num(r[3]),
                "act_hrs": num(r[4]), "dev_cost": num(r[5]), "ai_uses": num(r[6]),
                "ai_cost": num(r[7]), "total": num(r[8]),
                "notes": cell(r[9]) if len(r) > 9 else "",
            })
        if b == "TOTAL":
            cost_totals = {"weeks": num(r[2]), "est_hrs": num(r[3]), "act_hrs": num(r[4]),
                           "dev_cost": num(r[5]), "ai_uses": num(r[6]), "ai_cost": num(r[7]),
                           "total": num(r[8])}

    # ---- Platform Requirements coverage (from exec summary: 19/12/1/0) -------
    req_coverage = {"Covered": 19, "Planned": 12, "Partial": 1, "Gaps": 0}

    # ---- git state ----------------------------------------------------------
    def git(*args):
        try:
            return subprocess.run(["git", "-C", str(MONOREPO)] + list(args),
                                  capture_output=True, text=True, encoding="utf-8").stdout.strip()
        except Exception:
            return ""
    branch = git("rev-parse", "--abbrev-ref", "HEAD") or "—"
    log = git("log", "--pretty=format:%h|%ad|%s", "--date=short", "-n", "10")
    commits = []
    for line in log.splitlines():
        parts = line.split("|", 2)
        if len(parts) == 3:
            commits.append({"hash": parts[0], "date": parts[1], "subject": parts[2]})

    # ---- build phase model with live gate counts ---------------------------
    phase_model = []
    for pid, label, group, short, wstart, wend, prime, alpha, checkpoint in PHASES:
        g = gate_by_phase.get(pid, Counter())
        passed = g.get("Pass", 0) + g.get("Passed", 0)
        pending = sum(v for k, v in g.items() if k not in ("Pass", "Passed"))
        phase_model.append({
            "id": pid, "label": label, "group": group, "short": short,
            "wstart": wstart, "wend": wend,
            "prime": prime, "alpha": alpha, "checkpoint": checkpoint,
            "gates_total": passed + pending, "gates_passed": passed,
            "gaps": gaps_by_phase.get(pid, 0),
            "decisions": decisions_by_phase.get(pid, 0),
        })

    data = {
        "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "project": {
            "name": "Docmee",
            "tagline": "AI chatbot platform for medical clinics in Guatemala — WhatsApp-first, multi-tenant SaaS.",
            "status": "Sprint 0 — code built, largely uncommitted",
            "branch": branch,
            "plan_weeks": 26,
        },
        "headline": {
            "phases_total": 12, "phases_sealed": 12,
            "decisions": decisions_total, "decisions_flagged": flagged,
            "gates_total": sum(p["gates_total"] for p in phase_model),
            "gates_passed": sum(p["gates_passed"] for p in phase_model),
            "gaps_total": sum(gaps_by_phase.values()),
            "gaps_locked": gaps_status.get("Locked", 0) + gaps_status.get("Resolved", 0),
            "risks": len(risks),
            "security": len(security),
        },
        "phases": phase_model,
        "track_seed": TRACK_SEED,
        "sprint0": SPRINT0_TICKETS_as_dicts(),
        "agents": {
            "prime": {"name": "Prime", "role": "Backend", "packages": [{"path": p, "desc": d} for p, d in PRIME_PKGS]},
            "alpha": {"name": "Alpha", "role": "Frontend", "packages": [{"path": p, "desc": d} for p, d in ALPHA_PKGS]},
        },
        "tracker": {
            "actions": actions, "action_status": dict(action_status),
            "security": security, "sec_priority": dict(sec_priority), "sec_status": dict(sec_status_simple),
            "risks": risks, "risk_priority": dict(risk_priority),
            "gaps_status": dict(gaps_status), "gaps_total": sum(gaps_by_phase.values()),
            "decisions_total": decisions_total, "decisions_flagged": flagged,
            "future": future, "req_coverage": req_coverage,
            "gate_rows": gate_rows,
        },
        "git": {"branch": branch, "commits": commits},
        "cost": {"assumptions": assumptions, "items": cost_items, "totals": cost_totals},
    }
    return data

def main():
    data = build_data()
    tpl = TEMPLATE.read_text(encoding="utf-8")
    html = tpl.replace("/*__DOCMEE_DATA__*/null",
                       json.dumps(data, ensure_ascii=False))
    OUT.write_text(html, encoding="utf-8")
    h = data["headline"]
    print(f"Wrote {OUT}  ({len(html):,} bytes)")
    print(f"  phases={len(data['phases'])} gates={h['gates_total']} "
          f"actions={len(data['tracker']['actions'])} security={len(data['tracker']['security'])} "
          f"risks={len(data['tracker']['risks'])} decisions={h['decisions']} gaps={h['gaps_total']}")

def SPRINT0_TICKETS_as_dicts():
    return [{"id": t[0], "agent": t[1], "title": t[2], "detail": t[3], "status": t[4]}
            for t in SPRINT0_TICKETS]

if __name__ == "__main__":
    main()
