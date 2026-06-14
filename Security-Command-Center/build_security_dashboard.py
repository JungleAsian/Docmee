# -*- coding: utf-8 -*-
"""
Docmee Security Command Center - dashboard generator.

Reads the authored AI security analysis (findings.json), cross-checks it against
the live tracker (docmee-trackers.xlsx: Security Audit + Risk Register) for drift /
new items, scans the live code (monorepo + Backend + Frontend) with heuristic
patterns, pulls monorepo git state, and emits ONE self-contained HTML file
(Security-Dashboard.html) - opens offline by double-click, no server, no CDN.

This is the MECHANICAL refresh: it re-reads data + re-scans code every run. It does
NOT re-run the AI analysis - that lives in findings.json and is refreshed by a deep
re-audit (see README "Refresh"). Run hourly via the scheduled task, or on demand:

    python build_security_dashboard.py
"""
import json, subprocess, datetime, pathlib, sys, re
from collections import Counter

HERE     = pathlib.Path(__file__).resolve().parent
ROOT     = HERE.parent                       # ...\Docmee
FINDINGS = HERE / "findings.json"
TEMPLATE = HERE / "template.html"
OUT      = HERE / "Security-Dashboard.html"
STATE    = HERE / "scan-state.json"
XLSX     = ROOT / "Trackers" / "docmee-trackers.xlsx"
MONOREPO = ROOT / "monorepo"
SCAN_DIRS = [ROOT / "monorepo" / "apps", ROOT / "monorepo" / "packages",
             ROOT / "Backend", ROOT / "Frontend"]

SEV_RANK = {"Critical": 0, "High": 1, "Medium-High": 2, "Medium": 3, "Low": 4, "Info": 5}

# ---- heuristic code-scan patterns (low false-positive, high signal) ----------
CODE_EXT = {".ts", ".tsx", ".js", ".jsx", ".mjs", ".cjs", ".json", ".yaml", ".yml", ".env", ".sql"}
SCAN_PATTERNS = [
    ("Hard-coded secret literal", re.compile(r"""(?i)(api[_-]?key|secret|password|passwd|private[_-]?key)\s*[:=]\s*['"][^'"\s]{8,}['"]""")),
    ("Possible PHI in URL/query", re.compile(r"""(?i)(\?|&)(name|phone|email|dob|patient|clinic_id|ssn)=""")),
    ("clinic_id read from request", re.compile(r"""(?i)(req(uest)?\.(body|query|params)|searchParams\.get)\s*[.\[(]?\s*['"]?clinic_?id""")),
    ("dangerouslySetInnerHTML", re.compile(r"dangerouslySetInnerHTML")),
    ("eval() / new Function()", re.compile(r"\beval\s*\(|new\s+Function\s*\(")),
    ("child_process / exec", re.compile(r"child_process|\bexecSync?\s*\(|\bspawnSync?\s*\(")),
    ("CORS wildcard origin", re.compile(r"""(?i)(access-control-allow-origin|origin)\s*[:=]\s*['"]\*['"]""")),
    ("Non-HttpOnly cookie write (document.cookie)", re.compile(r"document\.cookie\s*=")),
    ("TLS verify disabled", re.compile(r"(?i)rejectUnauthorized\s*:\s*false|NODE_TLS_REJECT_UNAUTHORIZED")),
    ("Console logging (PHI risk - review)", re.compile(r"console\.(log|info|debug)\s*\(")),
]


def fail(msg):
    sys.stderr.write(msg + "\n")
    sys.exit(1)


def git(*args):
    try:
        return subprocess.run(["git", "-C", str(MONOREPO)] + list(args),
                              capture_output=True, text=True, encoding="utf-8").stdout.strip()
    except Exception:
        return ""


def load_tracker():
    """Return ({SECxx: (priority,status,concern)}, {Rxx: (priority,risk,status)}) or ({},{}) if unavailable."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX, data_only=True)
    except Exception as e:
        sys.stderr.write(f"[warn] tracker unavailable ({e}); drift-check skipped, using findings.json only\n")
        return {}, {}, False
    sec, risk = {}, {}
    def c(v): return "" if v is None else str(v).strip()
    if "Security Audit" in wb.sheetnames:
        for r in wb["Security Audit"].iter_rows(min_row=2, values_only=True):
            if r and r[0] and c(r[0]).startswith("SEC"):
                row = [c(x) for x in r] + [""] * 8
                sec[c(r[0])] = (row[5], row[7], row[2])
    if "Risk Register" in wb.sheetnames:
        for r in wb["Risk Register"].iter_rows(min_row=2, values_only=True):
            if r and r[0] and c(r[0]).startswith("R") and c(r[0])[1:2].isdigit():
                row = [c(x) for x in r] + [""] * 9
                risk[c(r[0])] = (row[6], row[1], row[8])
    return sec, risk, True


def scan_code():
    """Walk scan dirs, count heuristic pattern hits (excluding node_modules / build output)."""
    hits = {label: {"count": 0, "files": []} for label, _ in SCAN_PATTERNS}
    files_scanned = 0
    dir_file_counts = {}
    SKIP = {"node_modules", ".next", "dist", "build", ".git", "coverage", ".turbo"}
    for base in SCAN_DIRS:
        cnt = 0
        if not base.exists():
            dir_file_counts[base.name] = 0
            continue
        for p in base.rglob("*"):
            if not p.is_file() or p.suffix.lower() not in CODE_EXT:
                continue
            if any(part in SKIP for part in p.parts):
                continue
            cnt += 1
            files_scanned += 1
            try:
                text = p.read_text(encoding="utf-8", errors="ignore")
            except Exception:
                continue
            rel = str(p.relative_to(ROOT)).replace("\\", "/")
            for label, rx in SCAN_PATTERNS:
                if rx.search(text):
                    h = hits[label]
                    h["count"] += 1
                    if len(h["files"]) < 12 and rel not in h["files"]:
                        h["files"].append(rel)
        dir_file_counts[base.name] = cnt
    scan = [{"label": l, "count": hits[l]["count"], "files": hits[l]["files"]}
            for l, _ in SCAN_PATTERNS]
    return scan, files_scanned, dir_file_counts


def main():
    if not FINDINGS.exists():
        fail(f"missing {FINDINGS}")
    if not TEMPLATE.exists():
        fail(f"missing {TEMPLATE}")
    fdata = json.loads(FINDINGS.read_text(encoding="utf-8"))
    meta = fdata.get("meta", {})
    findings = list(fdata.get("findings", []))
    risk_overlay = fdata.get("risk_overlay", {})

    sec_live, risk_live, tracker_ok = load_tracker()

    # ---- synthesize lightweight risk findings from overlay + live tracker ----
    authored_ids = {f["id"] for f in findings}
    for rid, ov in risk_overlay.items():
        if rid in authored_ids:
            continue
        prio, title, status = risk_live.get(rid, ("Low", rid, ov.get("note", "")))
        findings.append({
            "id": rid, "source": "risk", "category": "Risk Register", "owner": "",
            "title": title or rid, "severity": _sev_from_priority(prio),
            "tracker_status": status or "", "assessment": ov.get("assessment", "holds"),
            "confidence": ov.get("confidence", 6), "phase": "",
            "threat": ov.get("note", ""), "blast": "", "cause": "",
            "solution": ov.get("note", ""), "action": "", "lightweight": True,
        })

    # ---- drift + new-item detection against the live tracker ----------------
    drift, unanalyzed = [], []
    if tracker_ok:
        for f in findings:
            fid = f["id"]
            # sec rows store status at idx 1; risk rows store status at idx 2.
            if fid in sec_live:
                live_status = sec_live[fid][1]
            elif fid in risk_live:
                live_status = risk_live[fid][2]
            else:
                continue
            if live_status and _norm(live_status) != _norm(f.get("tracker_status", "")):
                drift.append({"id": fid, "title": f["title"],
                              "was": f.get("tracker_status", ""), "now": live_status})
        known = {f["id"] for f in findings}
        for sid in sorted(sec_live):
            if sid not in known:
                unanalyzed.append({"id": sid, "title": sec_live[sid][2], "kind": "Security"})
        for rid in sorted(risk_live):
            if rid not in known:
                unanalyzed.append({"id": rid, "title": risk_live[rid][1], "kind": "Risk"})

    # ---- code scan + git ----------------------------------------------------
    scan, files_scanned, dir_counts = scan_code()
    code_hits_total = sum(s["count"] for s in scan if s["label"] != "Console logging (PHI risk - review)")
    branch = git("rev-parse", "--abbrev-ref", "HEAD") or "-"
    head = git("rev-parse", "--short", "HEAD") or "-"
    log = git("log", "--pretty=format:%h|%ad|%s", "--date=short", "-n", "8")
    commits = []
    for line in log.splitlines():
        parts = line.split("|", 2)
        if len(parts) == 3:
            commits.append({"hash": parts[0], "date": parts[1], "subject": parts[2]})

    audit_head = meta.get("audit_head", "")
    code_changed = bool(audit_head) and head != "-" and not head.startswith(audit_head[:7]) and not audit_head.startswith(head[:7])
    be_fe_files = dir_counts.get("Backend", 0) + dir_counts.get("Frontend", 0)

    # ---- sort + bucket ------------------------------------------------------
    findings.sort(key=lambda f: (SEV_RANK.get(f.get("severity", "Medium"), 3), -int(f.get("confidence", 0))))
    action_plan = [f for f in findings if f.get("assessment") == "action" and int(f.get("confidence", 0)) >= meta.get("action_threshold", 8)]
    watchlist = [f for f in findings if int(f.get("confidence", 0)) < meta.get("action_threshold", 8)]
    holds = [f for f in findings if f.get("assessment") == "holds" and int(f.get("confidence", 0)) >= meta.get("action_threshold", 8)]

    prio_counts = Counter(f.get("severity", "?") for f in findings)
    assess_counts = Counter(f.get("assessment", "?") for f in findings)
    conf_hist = Counter(int(f.get("confidence", 0)) for f in findings)

    open_sev = Counter(f.get("severity") for f in action_plan)

    data = {
        "generated": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "meta": meta,
        "project": {"name": "Docmee", "branch": branch, "head": head,
                    "tagline": "Independent security oversight for the Docmee build - findings, blast radius, and fixes rated by confidence."},
        "freshness": {
            "audit_date": meta.get("audit_date", ""), "audit_head": audit_head,
            "monorepo_head": head, "code_changed_since_audit": code_changed,
            "backend_files": dir_counts.get("Backend", 0), "frontend_files": dir_counts.get("Frontend", 0),
            "be_fe_files": be_fe_files, "files_scanned": files_scanned,
            "tracker_ok": tracker_ok,
        },
        "headline": {
            "total": len(findings),
            "action": len(action_plan),
            "actionable_hi_conf": len([f for f in action_plan]),
            "critical_high_open": open_sev.get("Critical", 0) + open_sev.get("High", 0),
            "code_hits": code_hits_total,
            "holds": len(holds),
            "watchlist": len(watchlist),
            "drift": len(drift), "unanalyzed": len(unanalyzed),
        },
        "prio_counts": dict(prio_counts),
        "assess_counts": dict(assess_counts),
        "conf_hist": {str(k): conf_hist.get(k, 0) for k in range(0, 11)},
        "findings": findings,
        "action_plan": action_plan,
        "watchlist": watchlist,
        "scan": scan,
        "dir_counts": dir_counts,
        "drift": drift,
        "unanalyzed": unanalyzed,
        "git": {"branch": branch, "commits": commits},
    }

    html = TEMPLATE.read_text(encoding="utf-8").replace(
        "/*__SECDATA__*/null", json.dumps(data, ensure_ascii=False))
    OUT.write_text(html, encoding="utf-8")
    STATE.write_text(json.dumps({
        "ts": data["generated"], "head": head, "branch": branch,
        "be_fe_files": be_fe_files, "code_hits": code_hits_total,
        "drift": len(drift), "unanalyzed": len(unanalyzed),
    }, indent=2), encoding="utf-8")

    print(f"Wrote {OUT}  ({len(html):,} bytes)")
    print(f"  findings={len(findings)} action_plan={len(action_plan)} watchlist={len(watchlist)} "
          f"code_hits={code_hits_total} scanned={files_scanned} drift={len(drift)} unanalyzed={len(unanalyzed)}")
    if code_changed:
        print(f"  [!] monorepo HEAD {head} != audit head {audit_head} - deep re-audit recommended.")
    if be_fe_files:
        print(f"  [!] Backend/Frontend now has {be_fe_files} files - deep re-audit recommended.")


def _norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())[:24]


def _sev_from_priority(p):
    p = (p or "").lower()
    if "med-high" in p or "medium-high" in p:
        return "Medium-High"
    if "high" in p or "crit" in p:
        return "High"
    if "med" in p:
        return "Medium"
    return "Low"


if __name__ == "__main__":
    main()
